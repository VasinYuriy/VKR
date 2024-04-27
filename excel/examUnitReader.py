from openpyxl import load_workbook


class SyllabusReader:
    def __init__(self, path):
        self.sheet = load_workbook(path)['План']

    def find_cell(self, value):
        for column in range(1, self.sheet.max_column+1):
            if self.sheet.cell(column=column, row=3).value == value:
                return column

    def read_syllabus(self):
        discs = {}
        pracs = {}
        gias = {}
        facs = {}
        current_dict = discs
        global_units = {}
        unit_sum = 0
        hour_sum = 0
        row = 4
        first_cell = self.sheet.cell(column=1, row=row).value
        unit_column = self.find_cell('Факт')
        hour_column = self.find_cell('Конт. раб.')
        while first_cell:

            if 'Блок 1' in first_cell:
                unit_sum += int(self.sheet.cell(column=unit_column, row=row).value)
                hour_sum += float(self.sheet.cell(column=hour_column, row=row).value)

            elif 'Блок 2' in first_cell:
                unit = int(self.sheet.cell(column=unit_column, row=row).value)
                hour_sum += float(self.sheet.cell(column=hour_column, row=row).value)
                unit_sum += unit
                global_units['Практики'] = unit
                current_dict = pracs

            elif 'Блок 3' in first_cell:
                unit = int(self.sheet.cell(column=unit_column, row=row).value)
                hour_sum += float(self.sheet.cell(column=hour_column, row=row).value)
                unit_sum += unit
                global_units['Государственная итоговая аттестация'] = unit
                current_dict = gias

            elif 'ФТД' in first_cell:
                global_units['Факультативные дисциплины'] = int(self.sheet.cell(column=unit_column, row=row).value)
                current_dict = facs

            exam_cell = self.sheet.cell(column=3, row=row)
            unit_cell = self.sheet.cell(column=unit_column, row=row)
            unit = unit_cell.value

            # if self.sheet.cell(column=2, row=row).value and '(У)' in self.sheet.cell(column=2, row=row).value:
            #     exam = 'Учебная ' + exam_cell.value[0].lower() + exam_cell.value[1:]
            #     if exam_cell.value and not exam_cell.font.bold:
            #         if unit:
            #             current_dict[exam] = int(unit)
            #         else:
            #             current_dict[exam] = 'x'
            #         if self.sheet.cell(column=6, row=row).value:
            #             diffs.append(exam)
            #
            # elif self.sheet.cell(column=2, row=row).value and '(П)' in self.sheet.cell(column=2, row=row).value:
            #     exam = 'Производственная ' + exam_cell.value[0].lower() + exam_cell.value[1:]
            #     if exam_cell.value and not exam_cell.font.bold:
            #         if unit:
            #             current_dict[exam] = int(unit)
            #         else:
            #             current_dict[exam] = 'x'
            #         if self.sheet.cell(column=6, row=row).value:
            #             diffs.append(exam)

            if exam_cell.value and not exam_cell.font.bold:
                exam = exam_cell.value
                if 'У' in self.sheet.cell(column=2, row=row).value:
                    exam = 'Учебная практика ({})'.format(' '.join([word for word in exam.lower().split(' ') if word not in ['учебная', 'производственная']]))
                elif 'П' in self.sheet.cell(column=2, row=row).value:
                    exam = 'Производственная практика ({})'.format(' '.join([word for word in exam.lower().split(' ') if word not in ['учебная', 'производственная']]))
                if unit:
                    current_dict[exam] = int(unit)
                else:
                    current_dict[exam] = 'x'

            row += 1
            first_cell = self.sheet.cell(column=1, row=row).value

        exam_units = {
            'Дисциплины': discs,
            'Практики': pracs,
            'Государственная итоговая аттестация': gias,
            'Факультативы': facs
        }

        global_units['Объем образовательной программы'] = unit_sum
        global_units[
            'в том числе объем контактной работы обучающихся во взаимодействии с преподавателем в академических часах:'] = round(
            hour_sum)

        return exam_units, global_units


if __name__ == '__main__':
    reader = SyllabusReader('plan.xlsx')
    print(reader.read_syllabus())

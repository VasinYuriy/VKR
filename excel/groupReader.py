import openpyxl as opxl
import os


class GroupReader:
    def __init__(self, path_list):
        self.path_list = path_list
        self.first_name = [10, 1]
        self.first_subject = [9, 2]
        self.block_row = 8
        self.marks_dict = {}
        self.course_projects = {}
        self.course_works = {}
        self.subject_list = []
        self.excel_count_dict = {}

    def get_names(self, sheet):
        for row in range(self.first_name[0], sheet.max_row+1):

            name = sheet.cell(row=row, column=self.first_name[-1]).value

            if name not in self.marks_dict.keys():
                self.marks_dict[name] = {}
                self.course_projects[name] = {}
                self.course_works[name] = {}
                self.excel_count_dict[name] = [0, 0, False]

            subject_dict, course_projects_dict, course_works_dict, mark_sum, mark_count, three_flag = self.get_subjects(sheet, row)

            for subject in subject_dict.items():
                self.marks_dict[name][subject[0]] = subject[1]
            for subject in course_projects_dict.items():
                self.course_projects[name][subject[0]] = subject[1]

            for subject in course_works_dict.items():
                self.course_works[name][subject[0]] = subject[1]

            self.excel_count_dict[name][0] += mark_sum
            self.excel_count_dict[name][1] += mark_count

            if three_flag:
                self.excel_count_dict[name][2] = three_flag

    def get_subjects(self, sheet, row):
        subject_dict = {}
        course_projects_dict = {}
        course_works_dict = {}
        mark_sum = 0
        mark_count = 0
        three_flag = False

        for column in range(self.first_subject[-1], sheet.max_column+1):
            if sheet.cell(row=self.block_row, column=column).value:
                block = sheet.cell(row=self.block_row, column=column).value

            subject = sheet.cell(row=self.first_subject[0], column=column).value
            mark = sheet.cell(row=row, column=column).value

            if block == 'КП':
                course_projects_dict[subject] = mark
            elif block == 'КР':
                course_works_dict[subject] = mark
            else:
                subject_dict[subject] = mark
                if subject not in self.subject_list:
                    self.subject_list.append(subject)

            if not three_flag:
                if mark == 3:
                    three_flag = True
                elif mark in [4, 5]:
                    mark_count += 1
                    mark_sum += int(mark)

        subject_dict['Производственная практика (преддипломная)'] = 'None'


        return subject_dict, course_projects_dict, course_works_dict, mark_sum, mark_count, three_flag

    def get_info(self):
        for path in self.path_list:
            wb = opxl.load_workbook(path)
            ws = wb.active
            self.get_names(ws)
        excel_dict = {}
        for k, v in self.excel_count_dict.items():
            if not v[2] and v[1] and v[0]/v[1] >= 4.75:
                excel_dict[k] = True
            else:
                excel_dict[k] = False
        form = ws.cell(row=1, column=1).value.split(':')[-1]
        faculty = ws.cell(row=2, column=1).value.split(':')[-1]
        group = ws.cell(row=3, column=1).value.split(':')[-1]
        code = ws.cell(row=4, column=1).value.split(':')[-1].split('_')[0]
        self.subject_list.append('Производственная практика (преддипломная)')
        group_info = {
            'form': form,
            'faculty': faculty,
            'group': group.strip(),
            'code': code,
            'group_marks': self.marks_dict,
            'course_projects': self.course_projects,
            'course_works': self.course_works,
            'subjects': self.subject_list,
            'excel_dict': excel_dict,
        }

        return group_info


if __name__ == '__main__':
    paths = ['1.xlsx', '2.xlsx', '3.xlsx', '4.xlsx', '5.xlsx', '6.xlsx', '7.xlsx', '8.xlsx']
    # paths = ['1.xlsx']
    group_path = 'Д-Э406'
    paths = [os.path.join(group_path, path) for path in paths]
    like = GroupReader(path_list=paths)
    info = (like.get_info())
    print(info)






